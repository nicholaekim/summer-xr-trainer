"""Write a recorded selection to an organized Excel workbook.

The wide one-row-per-frame CSV (189 columns) is unreadable. This produces a
tidy, filterable table instead: one row per (frame x hand x joint), with a
small summary sheet up front. World positions are used (they move with the
hand), alongside the raw rotation quaternion that actually drives the motion.
"""
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# One row per (frame, hand, joint). Keep in sync with the rows passed in.
DATA_HEADERS = [
    "frame", "elapsed_s", "clock", "hand", "joint",
    "world_x", "world_y", "world_z",
    "qw", "qx", "qy", "qz",
    "moved",
]

_HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_LEFT_FILL = PatternFill("solid", fgColor="E8FBFB")   # pale cyan
_RIGHT_FILL = PatternFill("solid", fgColor="FDECEC")  # pale red
_MOVED_FONT = Font(bold=True, color="9A6A00")
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER


def write_workbook(
    path: Path,
    *,
    summary: Sequence[tuple],
    rows: Iterable[Sequence],
) -> None:
    """summary: list of (label, value) pairs. rows: iterable matching DATA_HEADERS."""
    wb = Workbook()

    # --- Summary sheet ------------------------------------------------
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum["A1"] = "XR hand recording - selection export"
    ws_sum["A1"].font = Font(bold=True, size=14)
    r = 3
    for label, value in summary:
        ws_sum.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws_sum.cell(row=r, column=2, value=value)
        r += 1
    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 42

    # --- Joint data sheet (tidy long table) ---------------------------
    ws = wb.create_sheet("Joint data")
    ws.append(DATA_HEADERS)
    ncols = len(DATA_HEADERS)
    _style_header(ws, ncols)

    num_cols = {"world_x", "world_y", "world_z", "qw", "qx", "qy", "qz"}
    num_idx = {DATA_HEADERS.index(n) for n in num_cols}

    for row in rows:
        ws.append(list(row))
        rr = ws.max_row
        hand = row[DATA_HEADERS.index("hand")]
        moved = row[DATA_HEADERS.index("moved")]
        fill = _LEFT_FILL if hand == "left" else _RIGHT_FILL
        for c in range(1, ncols + 1):
            cell = ws.cell(row=rr, column=c)
            cell.fill = fill
            cell.border = _BORDER
            if (c - 1) in num_idx:
                cell.number_format = "0.000"
                cell.alignment = Alignment(horizontal="right")
        if moved == "yes":
            ws.cell(row=rr, column=DATA_HEADERS.index("joint") + 1).font = _MOVED_FONT

    # Frozen header + filter + sensible widths.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"
    widths = {
        "frame": 8, "elapsed_s": 11, "clock": 14, "hand": 7,
        "joint": 22, "moved": 7,
    }
    for i, name in enumerate(DATA_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 10)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
